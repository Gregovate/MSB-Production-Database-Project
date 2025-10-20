#!/usr/bin/env python3
"""
merge_reports_to_excel.py
--------------------------------
Baseline, self-contained script to combine merger CSVs into a single Excel file
with multiple tabs, frozen headers, filters, auto-width, and simple coloring.

Usage (Windows PowerShell):
    python merge_reports_to_excel.py

Dependencies:
    pip install pandas openpyxl
"""

from pathlib import Path
from datetime import datetime
import warnings
import pandas as pd
import getpass
import platform
import argparse
# GAL 25-10-15: needed for regex in _contains_any
import re

from pandas.errors import ParserWarning
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill

# Silence noisy CSV parser warnings (optional)
warnings.simplefilter("ignore", ParserWarning)

# ================== CONFIG ==================
import argparse
from pathlib import Path
from datetime import datetime

parser = argparse.ArgumentParser(description="Merge CSV reports to a single Excel.")
parser.add_argument(
    "--root",
    default=r"G:\Shared drives\MSB Database\Database Previews\reports",  # CSV input folder
    help="Folder containing the input CSVs (compare, ledger, etc.).",
)
parser.add_argument(
    "-o", "--out",
    default=r"G:\Shared drives\MSB Database\Database Previews",          # Excel output folder
    help="Folder to write the Excel into.",
)
args = parser.parse_args()

ROOT = Path(args.root)                 # where CSVs are read from
OUT_DIR = Path(args.out)               # where Excel is written
OUT_DIR.mkdir(parents=True, exist_ok=True)

# Which CSVs → which sheet names
FILES = [
    #("compare.csv", "Compare"),
    ("missing_comments.csv", "Missing_Comments"),
    ("all_staged_comments.csv", "All_Staged_Comments"),
    ("excluded_winners.csv", "Excluded_Winners"),
    ("applied_this_run.csv", "Applied_This_Run"),
    ("apply_events.csv", "Apply_Events"),
    ("current_previews_ledger.csv", "Current_Previews_Ledger"),
    ("revision_mismatches.csv", "Revision_Mismatches"),
    ("needs_action.csv", "NeedsAction"),   # <<< NEW: prefer the CSV we generate in dry-run
]

STAMP = datetime.now().strftime("%Y%m%d-%H%M")
OUT_XLSX_STAMPED = OUT_DIR / f"lorprev_reports-{STAMP}.xlsx"
OUT_XLSX_FIXED   = OUT_DIR / "lorprev_reports.xlsx"

print(f"[cfg] CSV root: {ROOT}")
print(f"[cfg] Excel out: {OUT_DIR}")
# ============================================





def read_csv_safe(path: Path) -> pd.DataFrame:
    """
    Simple robust CSV reader:
    - UTF-8 with BOM support
    - engine='python' with delimiter sniffing (sep=None)
    - on_bad_lines='skip' to avoid fatal parse errors
    """
    if not path.exists():
        return pd.DataFrame()
    return pd.read_csv(
        path,
        dtype=str,
        keep_default_na=False,
        encoding="utf-8-sig",
        sep=None,            # auto-detect delimiter
        engine="python",
        on_bad_lines="skip", # skip malformed rows instead of crashing
        quoting=0            # QUOTE_MINIMAL
    )

# ---------------------------------------------------------------------------
# GAL 25-10-15: Robust NeedsAction builder + ledger annotation
# ---------------------------------------------------------------------------
NEEDS_TOKENS = {
    # Any of these words appearing in action/status/reason should qualify
    "any": (
        "update-staging", "stage-new", "ready to apply",
        "out-of-date", "replace", "apply", "needs action",
        "winner newer", "newer export", "semantic different",
        "not staged", "missing in staging",
    )
}

def _norm(s):  return "" if s is None else str(s).strip()
def _low(s):   return _norm(s).lower()

def _pick_key_cols(df):
    key_col  = "Key" if "Key" in df.columns else None
    name_col = "PreviewName" if "PreviewName" in df.columns else None
    return key_col, name_col

def _contains_any(series: pd.Series, needles: tuple[str, ...]) -> pd.Series:
    if series is None:
        return pd.Series(False, index=[])
    s = series.astype(str).str.lower()
    # use regex OR of all needles, escape spaces with \s* around hyphens
    pattern = "|".join([re.escape(x) for x in needles])
    return s.str.contains(pattern, regex=True, na=False)

def _describe_action(row: dict) -> str:
    a = _low(row.get("Action"))
    r = _low(row.get("Reason"))
    st = _low(row.get("Status") or row.get("Result") or row.get("Decision") or row.get("Outcome") or row.get("Comment"))

    if "stage-new" in a or "not staged" in r or "missing in staging" in r:
        return "Stage new"
    if "update-staging" in a or "replace" in a or "winner newer" in r or "newer export" in r:
        return "Replace staged (winner newer)"
    if "out-of-date" in a or "out-of-date" in st:
        return "Staged out-of-date"
    if "ready to apply" in a or "ready to apply" in st:
        return "Apply to staging"
    if "semantic different" in r:
        return "Replace staged (content changed)"
    # fallbacks
    if a:
        return a
    if st:
        return st
    return (row.get("Reason") or "Needs action").strip()

def _slice_common_cols(df, extra_cols=()):
    cols = ["PreviewName","Key","Revision","Exported","User","Action","Reason","Comment","Status","Path","StagedPath"]
    cols = [c for c in cols + list(extra_cols) if c in df.columns]
    return df.loc[:, cols].copy()

def build_needs_action_df(tables: dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows = []

    # 1) Compare: usually has the clearest apply actions # GAL 25-10-18: Compare removed — no data from compare.csv
    # cmp_df = tables.get("Compare")
    # if cmp_df is not None and not cmp_df.empty:
    #     act = cmp_df["Action"] if "Action" in cmp_df.columns else None
    #     reas = cmp_df["Reason"] if "Reason" in cmp_df.columns else None
    #     mask = _contains_any(act, NEEDS_TOKENS["any"]) | _contains_any(reas, NEEDS_TOKENS["any"])
    #     if mask.any():
    #         part = _slice_common_cols(cmp_df.loc[mask])
    #         rows.append(part)

    # 2) Staged (linked): pick out-of-date
    staged = tables.get("Staged")
    if staged is not None and not staged.empty:
        act = staged["Action"] if "Action" in staged.columns else None
        mask = _contains_any(act, ("out-of-date",))
        if mask.any():
            part = _slice_common_cols(staged.loc[mask])
            rows.append(part)

    # 3) Manifest (if exported): some pipelines write a manifest with Role/Action
    for key in ("Manifest", "Current_Previews_Manifest", "current_previews_manifest"):
        mf = tables.get(key)
        if mf is not None and not mf.empty:
            # look across Role/Action/Reason
            mask = pd.Series(False, index=mf.index)
            for col in ("Role","Action","Reason","WinnerReason"):
                if col in mf.columns:
                    mask = mask | _contains_any(mf[col], NEEDS_TOKENS["any"])
            if mask.any():
                part = _slice_common_cols(mf.loc[mask])
                rows.append(part)
            break

    # 4) Current_Previews_Ledger: scan status/comment style columns
    led = tables.get("Current_Previews_Ledger")
    if led is not None and not led.empty:
        # choose any column with status-ish naming
        cand_cols = [c for c in led.columns if any(t in c.lower() for t in ("status","action","result","decision","outcome","comment","reason"))]
        mask = pd.Series(False, index=led.index)
        for c in cand_cols:
            mask = mask | _contains_any(led[c], NEEDS_TOKENS["any"])
        if mask.any():
            part = _slice_common_cols(led.loc[mask], extra_cols=tuple(cand_cols))
            rows.append(part)

    if not rows:
        return pd.DataFrame()

    needs = pd.concat(rows, ignore_index=True)

    # Normalize and synthesize ActionNeeded
    for c in ("PreviewName","Key","Revision","Exported","User","Action","Reason","Comment","Status","Path","StagedPath"):
        if c in needs.columns:
            needs[c] = needs[c].map(_norm)

    needs["ActionNeeded"] = needs.apply(lambda r: _describe_action(r.to_dict()), axis=1)

    # Deduplicate by Key then PreviewName
    key_col, name_col = _pick_key_cols(needs)
    id_col = key_col or name_col
    if id_col:
        # prefer higher Revision, then newer Exported if present
        def _rev_int(x):
            try: return int(str(x).strip())
            except: return -1
        def _to_dt(x):
            from datetime import datetime
            s = str(x)
            for fmt in ("%Y-%m-%d %H:%M:%S","%Y-%m-%d %H:%M","%m/%d/%Y %H:%M","%Y-%m-%d"):
                try: return datetime.strptime(s, fmt)
                except: pass
            return None
        needs["_rev"] = needs.get("Revision", "").map(_rev_int) if "Revision" in needs.columns else -1
        needs["_exp"] = needs.get("Exported", "").map(_to_dt) if "Exported" in needs.columns else None
        needs = (
            needs.sort_values(by=[id_col, "_rev", "_exp"], ascending=[True, False, False], na_position="last")
                 .drop_duplicates(subset=[id_col], keep="first")
        )
        if "_rev" in needs.columns: needs.drop(columns=["_rev"], inplace=True)
        if "_exp" in needs.columns: needs.drop(columns=["_exp"], inplace=True)

    # Final column order
    ordered = [c for c in ("PreviewName","Key","Revision","Exported","User","ActionNeeded","Action","Reason","Comment","Path","StagedPath") if c in needs.columns]
    needs = needs[ordered].reset_index(drop=True)
    return needs

def annotate_ledger_with_needs_action(tables: dict[str, pd.DataFrame], needs: pd.DataFrame) -> None:
    df = tables.get("Current_Previews_Ledger")
    if df is None or df.empty or needs is None or needs.empty:
        return
    ledger = df.copy()
    key_col, name_col = _pick_key_cols(ledger)
    if not key_col and not name_col:
        tables["Current_Previews_Ledger"] = ledger
        return

    # Map id -> ActionNeeded
    by_id = {}
    if key_col and key_col in needs.columns:
        for i, row in needs.iterrows():
            by_id[_norm(row[key_col])] = row.get("ActionNeeded","Needs action")
    elif name_col and name_col in needs.columns:
        for i, row in needs.iterrows():
            by_id[_norm(row[name_col])] = row.get("ActionNeeded","Needs action")

    # Ensure Comment column exists
    if "Comment" not in ledger.columns:
        ledger["Comment"] = ""

    # Apply annotations
    if key_col and key_col in ledger.columns:
        ids = ledger[key_col].map(_norm)
    else:
        ids = ledger[name_col].map(_norm)

    def _merge_comment(idx, old):
        aid = by_id.get(_norm(ids.iloc[idx]))
        if not aid:
            return old
        if not old:
            return f"Needs action — {aid}"
        low = old.lower()
        if "needs action" in low:
            return old  # don't duplicate
        return f"{old}; Needs action — {aid}"

    ledger["Comment"] = [ _merge_comment(i, _norm(v)) for i, v in enumerate(ledger["Comment"]) ]
    tables["Current_Previews_Ledger"] = ledger
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# GAL 25-10-15: Excel formatting helpers (openpyxl)
# ---------------------------------------------------------------------------
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def _header_map(ws, header_row: int = 1) -> dict[str, str]:
    """Map normalized header text -> column letter."""
    m = {}
    for cell in ws[header_row]:
        if cell.value:
            m[str(cell.value).strip().lower()] = cell.column_letter
    return m

def _col_letter_for(ws, header_name: str, header_row: int = 1) -> str | None:
    """Find column letter by header (case-insensitive)."""
    hmap = _header_map(ws, header_row)
    return hmap.get(str(header_name).strip().lower())

def _number_format_int(ws, headers: list[str], header_row: int = 1) -> None:
    """
    GAL 25-10-15
    Format given header columns as integers (format '0').
    Ignores missing columns gracefully.
    """
    for h in headers:
        col = _col_letter_for(ws, h, header_row)
        if not col:
            continue
        for r in range(header_row + 1, ws.max_row + 1):
            c = ws[f"{col}{r}"]
            # only set format if the cell looks numeric
            try:
                if c.value is None:
                    continue
                # coerce strings like "42" to number visually
                if isinstance(c.value, str) and c.value.isdigit():
                    c.value = int(c.value)
                if isinstance(c.value, (int, float)):
                    c.number_format = "0"
                    c.alignment = Alignment(horizontal="right")
            except Exception:
                pass

def _number_format_float(ws, headers: list[str], header_row: int = 1, decimals: int = 2) -> None:
    """Format columns as floats with a fixed number of decimals."""
    fmt = "0." + "0" * max(0, decimals)
    for h in headers:
        col = _col_letter_for(ws, h, header_row)
        if not col:
            continue
        for r in range(header_row + 1, ws.max_row + 1):
            c = ws[f"{col}{r}"]
            try:
                if c.value is None:
                    continue
                if isinstance(c.value, str):
                    try:
                        c.value = float(c.value)
                    except Exception:
                        continue
                if isinstance(c.value, (int, float)):
                    c.number_format = fmt
                    c.alignment = Alignment(horizontal="right")
            except Exception:
                pass

def _number_format_datetime(ws, headers: list[str], header_row: int = 1) -> None:
    """
    Format date/time looking columns as 'yyyy-mm-dd hh:mm'.
    Leaves text that can’t be parsed.
    """
    fmt = "yyyy-mm-dd hh:mm"
    # quick parser without external deps
    from datetime import datetime
    fmts_try = ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M",
                "%m/%d/%Y %H:%M", "%Y-%m-%d")
    for h in headers:
        col = _col_letter_for(ws, h, header_row)
        if not col:
            continue
        for r in range(header_row + 1, ws.max_row + 1):
            c = ws[f"{col}{r}"]
            v = c.value
            try:
                if v is None:
                    continue
                if isinstance(v, str):
                    dt = None
                    for f in fmts_try:
                        try:
                            dt = datetime.strptime(v.strip(), f)
                            break
                        except Exception:
                            pass
                    if dt is None:
                        continue
                    c.value = dt
                # openpyxl will handle python datetime types
                c.number_format = fmt
                c.alignment = Alignment(horizontal="left")
            except Exception:
                pass

def _auto_width(ws, header_row: int = 1, max_width: int = 60) -> None:
    """Autosize columns based on header + cell content length."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        try:
            lengths = [len(str(c.value)) if c.value is not None else 0 for c in col_cells]
            width = min(max_width, max(lengths + [10]) + 2)
            ws.column_dimensions[col_letter].width = width
        except Exception:
            pass


def autosize_and_filter(ws):
    """Freeze header, add autofilter, and auto-fit columns (capped width)."""
    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = 0
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 80)

def _choose_action_like_column(ws, header_row=1):
    """
    GAL 25-10-15: Prefer deterministic columns; fall back to heuristic.
    """
    title = ws.title
    # Prefer the mapped column if present
    preferred = STATUS_COLUMN_MAP.get(title)
    if preferred:
        # Find exact header match (case-insensitive)
        for cell in ws[header_row]:
            if cell.value and str(cell.value).strip().lower() == preferred.lower():
                return cell.column_letter, str(cell.value).strip()

    # Heuristic fallback (previous behavior)
    headers = [(cell.column_letter, str(cell.value).strip())
               for cell in ws[header_row] if cell.value]
    norm = [(col, txt, txt.lower()) for col, txt in headers]
    candidates = [(c, t, tl) for c, t, tl in norm if "date" not in tl and "time" not in tl]

    # Upgrade priority to include ActionNeeded
    priority = ("actionneeded", "status", "action", "result", "decision", "outcome", "operation", "comment")

    def score_col(letter):
        hits = 0
        max_row = min(ws.max_row, 400)
        for r in range(2, max_row + 1):
            v = ws[f"{letter}{r}"].value
            if not v:
                continue
            s = str(v).lower()
            if any(t in s for t in ("applied", "update-staging", "stage-new", "ready to apply", "out-of-date",
                                     "blocked", "error", "fail", "excluded", "skip", "identical", "no-op")):
                hits += 1
        return hits

    for key in priority:
        for col, txt, tl in candidates:
            if key in tl:
                return col, txt

    scored = [(score_col(col), col, txt) for col, txt, _ in candidates]
    if scored:
        scored.sort(reverse=True)
        if scored[0][0] > 0:
            return scored[0][1], scored[0][2]

    return None, None


def add_action_colors(ws, header_row=1):
    """Apply conditional colors to the most action/status-like column on this sheet."""
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import PatternFill

    col_letter, header_text = _choose_action_like_column(ws, header_row=header_row)
    if not col_letter or ws.max_row <= header_row:
        return

    rng = f"${col_letter}${header_row+1}:${col_letter}${ws.max_row}"
    top_left = f"{col_letter}{header_row+1}"  # e.g., D2 (row-relative)

    def any_of(cell_ref, terms):
        parts = [f'ISNUMBER(SEARCH("{t}",{cell_ref}))' for t in terms]
        return f"OR({','.join(parts)})"

    # Rules
    green  = any_of(top_left, ("applied", "update-staging", "allow", "allowed", "pass", "passed"))
    yellow = any_of(top_left, ("ready to apply",))
    red    = any_of(top_left, ("blocked", "error", "fail", "failed", "exclude", "excluded", "work needed"))
    gray   = any_of(top_left, ("skip", "identical", "no-op", "not needed", "unchanged"))

    for formula, color in (
        (green,  "C6EFCE"),   # light green
        (yellow, "FFF2CC"),   # light yellow
        (red,    "FFC7CE"),   # light red
        (gray,   "E7E6E6"),   # light gray
    ):
        ws.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[formula],
                fill=PatternFill(start_color=color, end_color=color, fill_type="solid"),
            ),
        )

    # Optional: keep this if you like the console hint
    print(f"[format] {ws.title}: using '{header_text}' column for color rules")

def add_missing_comments_colors(ws, header_row=1):
    """
    Highlight status-like columns in Missing_Comments:
    - Red if cell contains: needs / missing / blocked
    - Green if cell contains: ok / clean / complete
    Candidate columns: headers containing status/need/missing/reason/comment
    """
    # Identify candidate columns by header keywords
    status_cols = []
    for cell in ws[header_row]:
        if not cell.value:
            continue
        name = str(cell.value).strip().lower()
        if any(key in name for key in ("status", "need", "missing", "reason", "comment")):
            status_cols.append(cell.column_letter)

    if not status_cols or ws.max_row <= header_row:
        return

    last = ws.max_row
    for col_letter in status_cols:
        rng = f"${col_letter}${header_row+1}:${col_letter}${last}"
        # Red for needs/missing/blocked
        ws.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[f'OR(ISNUMBER(SEARCH("needs",$${col_letter}{header_row+1})),ISNUMBER(SEARCH("missing",$${col_letter}{header_row+1})),ISNUMBER(SEARCH("blocked",$${col_letter}{header_row+1})))'.replace("$$","$")],
                fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            )
        )
        # Green for ok/clean/complete
        ws.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[f'OR(ISNUMBER(SEARCH("ok",$${col_letter}{header_row+1})),ISNUMBER(SEARCH("clean",$${col_letter}{header_row+1})),ISNUMBER(SEARCH("complete",$${col_letter}{header_row+1})))'.replace("$$","$")],
                fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            )
        )

# ---------------------------------------------------------------------------
# GAL 25-10-15: Deterministic Overview — fixed status columns per sheet
# ---------------------------------------------------------------------------

# Which column to summarize per sheet (skip if sheet not present or column missing)
STATUS_COLUMN_MAP = {
    "Compare": "Action",
    "Staged": "Action",
    "Applied_This_Run": "Action",
    "Excluded_Winners": "Reason",
    "Missing_Comments": "Comment",
    "Current_Previews_Ledger": "Comment",
    "Revision_Mismatches": None,          # no status roll-up
    "All_Staged_Comments": None,          # very verbose
    "NeedsAction": "ActionNeeded",        # synthesized action we created
}

def _normalize_status_value(raw: str) -> str:
    """GAL 25-10-15: collapse noisy text into consistent buckets for Overview."""
    s = ("" if raw is None else str(raw)).strip().lower()
    if not s or s == "(blank)":
        return "(blank)"

    # common buckets
    if any(t in s for t in ("applied", "applies ok")):
        return "applied"
    if any(t in s for t in ("update-staging", "replace staged", "winner newer", "semantic different")):
        return "update-staging"
    if "stage-new" in s or "not staged" in s or "missing in staging" in s:
        return "stage-new"
    if "ready to apply" in s:
        return "ready to apply"
    if "out-of-date" in s:
        return "out-of-date"
    if any(t in s for t in ("blocked", "error", "fail", "failed", "exclude", "excluded", "work needed")):
        return "blocked/error"
    if any(t in s for t in ("skip", "identical", "no-op", "unchanged", "not needed")):
        return "skip/identical"
    # otherwise keep a short form (first 40 chars) so we don’t flood the overview
    return s[:40]

def build_overview(tables: dict[str, pd.DataFrame]) -> pd.DataFrame:
    """
    GAL 25-10-15
    Build a compact overview by counting normalized status values in a
    deterministic column per sheet (see STATUS_COLUMN_MAP).
    """
    rows = []
    for sheet, df in tables.items():
        col = STATUS_COLUMN_MAP.get(sheet)
        if not col:
            continue
        if col not in df.columns or df.empty:
            continue

        series = df[col].astype(str).replace({"": "(blank)"})
        counts = (
            series.map(_normalize_status_value)
                  .value_counts(dropna=False)
                  .rename_axis("Value")
                  .reset_index(name="Count")
        )
        for _, r in counts.iterrows():
            rows.append({
                "Sheet": sheet,
                "StatusColumn": col,
                "Value": str(r["Value"]),
                "Count": int(r["Count"]),
            })

    if not rows:
        return pd.DataFrame({"Info": ["No status columns found to summarize."]})

    return (
        pd.DataFrame(rows)
          .sort_values(["Sheet", "Count", "Value"], ascending=[True, False, True], ignore_index=True)
    )


# GAL 25-10-20: Restore User/Machine/Actor + optional Reason via run_meta.json
from datetime import datetime
from pathlib import Path
import os, json, getpass, platform, socket
import pandas as pd

def _who_am_i():
    user = os.getenv("USERNAME") or getpass.getuser() or "unknown"
    host = os.getenv("COMPUTERNAME") or platform.node() or socket.gethostname() or "unknown"
    return user, host, f"{user}@{host}"

def _load_run_meta(candidates):
    for base in candidates:
        try:
            if not base:
                continue
            meta_path = Path(base) / "run_meta.json"
            if meta_path.exists():
                return json.loads(meta_path.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}

def write_info_tab(
    writer,
    tables=None,
    root=None,
    compare_summary=None,
    run_mode=None,
    actor=None
):
    env_user, env_host, env_actor = _who_am_i()

    # Try run_meta.json in --root and CWD
    run_meta = _load_run_meta([root, Path.cwd()])
    meta_reason  = run_meta.get("reason") or ""
    meta_runmode = run_meta.get("run_mode") or run_meta.get("runMode") or ""
    meta_user    = run_meta.get("user") or env_user
    meta_host    = run_meta.get("host") or env_host
    meta_actor   = run_meta.get("actor") or f"{meta_user}@{meta_host}"

    info_row = {
        "Generated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "RunMode":   (run_mode or meta_runmode or ""),
        "Reason":    meta_reason,
        "User":      meta_user,
        "Machine":   meta_host,
        "Actor":     (actor or meta_actor or env_actor),
    }

    if compare_summary:
        for k, v in compare_summary.items():
            if k not in info_row:
                info_row[k] = v

    # NeedsAction count
    try:
        needs_df = None
        if isinstance(tables, dict) and "NeedsAction" in tables:
            needs_df = tables["NeedsAction"]
        elif root:
            p = Path(root) / "needs_action.csv"
            if p.exists():
                needs_df = pd.read_csv(p)
        if needs_df is not None:
            info_row["NeedsActionRows"] = int(len(needs_df))
    except Exception as e:
        print(f"[WARN] Info tab: could not read NeedsAction: {e}")

    # Optional: ledger count if provided in tables
    try:
        if isinstance(tables, dict):
            for key in ("Current_Previews_Ledger", "Ledger", "current_ledger"):
                if key in tables:
                    info_row["LedgerRows"] = int(len(tables[key]))
                    break
    except Exception as e:
        print(f"[WARN] Info tab: could not compute LedgerRows: {e}")

    pd.DataFrame([info_row]).to_excel(writer, index=False, sheet_name="Info")
    print("[GAL 25-10-20] Info tab written with RunMode/Reason/User/Machine/Actor")

def main():
    tables = {}
    for filename, sheet in FILES:
        df = read_csv_safe(ROOT / filename)
        if sheet == "Revision_Mismatches":
            for col in ("UsedRevision", "DiskLatestRevision"):
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors="coerce").astype("Int64")  # nullable int, no decimals
        if df is not None and not df.empty:
            tables[sheet] = df

    if not tables:
        print("[ERROR] No report CSVs found in:", ROOT)
        return

    # -----------------------------------------------------------------------
    # GAL 25-10-18: Prefer CSV NeedsAction (authoritative). If missing, fall back.
    # -----------------------------------------------------------------------
    def _ensure_actionneeded(df: pd.DataFrame) -> pd.DataFrame:
        if df is None or df.empty:
            return df
        if "ActionNeeded" not in df.columns:
            # Synthesize from ReadyToApply / Blockers for Overview coloring
            def _mk(row):
                ready = str(row.get("ReadyToApply","")).strip().lower()
                blocks = str(row.get("Blockers","")).strip().lower()
                if ready == "yes":
                    return "ready to apply"
                if blocks:
                    return "blocked"
                return "needs action"
            df = df.copy()
            df["ActionNeeded"] = df.apply(_mk, axis=1)
        return df

    needs = None
    if "NeedsAction" in tables and not tables["NeedsAction"].empty:
        needs = _ensure_actionneeded(tables["NeedsAction"])
    else:
        # Very old behavior (heuristic) as fallback only
        needs = build_needs_action_df(tables)
        needs = _ensure_actionneeded(needs)
        if needs is not None and not needs.empty:
            tables["NeedsAction"] = needs

    if needs is not None and not needs.empty:
        tables["NeedsAction"] = needs
        annotate_ledger_with_needs_action(tables, needs)
        print(f"[info][GAL 25-10-18] NeedsAction rows: {len(needs)} (from needs_action.csv)")
    else:
        print("[info][GAL 25-10-18] No NeedsAction rows found")



    with pd.ExcelWriter(OUT_XLSX_STAMPED, engine="openpyxl") as writer:
        # ---- Write Overview + Info FIRST so they appear at the front ----
        overview = build_overview(tables)
        overview.to_excel(writer, index=False, sheet_name="Overview")
        write_info_tab(
            writer,
            tables=tables,                 # your dict of DataFrames used for other sheets
            root=args.root,                # the --root path you already parse
            compare_summary=compare_summary if 'compare_summary' in locals() else None,
            run_mode=run_mode if 'run_mode' in locals() else None,
            actor=actor if 'actor' in locals() else None
)


        # ---- Now write the normal report tabs ----
        for sheet_name, df in tables.items():
            df.to_excel(writer, index=False, sheet_name=sheet_name)

        # ---- Post-format all sheets (keep your existing helpers/calls) ----
        wb = writer.book
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            autosize_and_filter(ws)
            if sheet_name not in ("Overview", "Info"):
                add_action_colors(ws)
            if sheet_name == "Missing_Comments":
                add_missing_comments_colors(ws)
        
            if sheet_name == "Revision_Mismatches":
                _number_format_int(ws, ["UsedRevision", "DiskLatestRevision"])

            # NeedsAction sheet: datetime + size formatting
            if sheet_name == "NeedsAction":
                _number_format_datetime(ws, ["AuthorFileTime", "StagedFileTime"])
                _number_format_int(ws, ["AuthorFileSize", "StagedFileSize"])

        # Make Overview the active sheet when the workbook opens
        if "Overview" in wb.sheetnames:
            wb.active = wb.sheetnames.index("Overview")


    print(f"[OK] Wrote Excel (timestamped): {OUT_XLSX_STAMPED}")

    # Best-effort copy to fixed name (skip if locked/open)
    try:
        OUT_XLSX_FIXED.write_bytes(OUT_XLSX_STAMPED.read_bytes())
        print(f"[OK] Also wrote: {OUT_XLSX_FIXED}")
    except PermissionError:
        print(f"[WARN] Could not overwrite {OUT_XLSX_FIXED} (file in use). Using timestamped output.")

if __name__ == "__main__":
    main()
