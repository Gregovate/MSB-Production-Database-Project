# GAL 2025-11-06  V1.3.0
# python apply_rgbplus_tags_from_excel.py
#
# Apply edited ApplyTag / ApplyPropTag + MaxChannels values from *_lor_groups.xlsx
# back into a .lorprev file.
#
# Typical usage (no arguments):
#   python apply_rgbplus_tags_from_excel.py
#   → file picker for the Excel file
#   → automatically locates the source .lorprev using Meta sheet
#   → writes retagged .lorprev into <excel_folder>\retagged_previews\
#
# Safety features:
#   - Meta sheet: ensures Excel belongs to the selected .lorprev
#   - Validates ApplyTag / ApplyPropTag values against previewtags.txt
#   - Uses PropID (Props) and MemberPropID (GroupMembers), merged
#   - Detects conflicting tag assignments for the same PropID
#
# Requires: openpyxl

from pathlib import Path
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
import argparse
import sys
import os
import tkinter as tk
from tkinter import filedialog


# ---------------------------------------------------------------------
# Tag source handling (LOR install + back door copy)
# ---------------------------------------------------------------------


def load_allowed_tags_from_lor(lorprev_path: Path):
    r"""
    Load allowed tags from previewtags.txt.

    Primary:  C:\lor\LORInternal\Downloads\previews\previewtags.txt
    Back door: <lorprev_folder>\previewtags.txt

    Returns a set of tags (may be empty).
    """
    default_path = Path(r"C:\lor\LORInternal\Downloads\previews\previewtags.txt")

    override = os.environ.get("LOR_PREVIEW_TAGS_PATH")
    tag_file = Path(override) if override else default_path

    tags = set()

    if tag_file.is_file():
        with tag_file.open("r", encoding="utf-8") as f:
            for line in f:
                t = line.strip()
                if t:
                    tags.add(t)
        return tags

    # Fallback in preview folder
    fallback = lorprev_path.parent / "previewtags.txt"
    if fallback.is_file():
        with fallback.open("r", encoding="utf-8") as f:
            for line in f:
                t = line.strip()
                if t:
                    tags.add(t)
        return tags

    return tags


# ---------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------


def load_meta(ws_meta):
    meta = {}
    for row in ws_meta.iter_rows(min_row=2, values_only=True):
        key, value = row[0], row[1]
        if key:
            meta[str(key)] = str(value) if value is not None else ""
    return meta


def load_apply_map_groups(ws_gm, allowed_set):
    """
    Build MemberPropID -> ApplyTag map from GroupMembers sheet.
    Returns (map, conflicts, invalid_tags).
    """
    header = [c.value for c in next(ws_gm.iter_rows(min_row=1, max_row=1))]
    try:
        idx_id = header.index("MemberPropID")
        idx_apply = header.index("ApplyTag")
    except ValueError:
        # Sheet present but missing columns; treat as empty
        return {}, [], []

    apply_map = {}
    conflicts = []
    invalid_tags = []

    for row in ws_gm.iter_rows(min_row=2, values_only=True):
        prop_id = str(row[idx_id]) if row[idx_id] is not None else ""
        apply_tag = row[idx_apply]
        if not prop_id:
            continue
        apply_tag = str(apply_tag).strip() if apply_tag is not None else ""
        if not apply_tag:
            continue

        # Validate against allowed tags if we have them
        if allowed_set and apply_tag not in allowed_set:
            invalid_tags.append(f"{prop_id}: {apply_tag}")

        if prop_id not in apply_map:
            apply_map[pid := prop_id] = apply_tag
        else:
            if apply_map[prop_id] != apply_tag:
                conflicts.append(
                    f"Conflicting ApplyTag for MemberPropID {prop_id}: "
                    f"{apply_map[prop_id]!r} vs {apply_tag!r}"
                )

    return apply_map, conflicts, invalid_tags


def load_apply_map_props(ws_props, allowed_set):
    """
    Build:
      - PropID -> ApplyPropTag map
      - PropID -> MaxChannels map (if provided)

    Returns (tag_map, invalid_tags_for_props, maxchannels_map).
    """
    header = [c.value for c in next(ws_props.iter_rows(min_row=1, max_row=1))]
    try:
        idx_id = header.index("PropID")
        idx_apply = header.index("ApplyPropTag")
    except ValueError:
        # Sheet present but missing columns; treat as empty
        return {}, [], {}

    idx_max = header.index("MaxChannels") if "MaxChannels" in header else None

    tag_map = {}
    invalid_tags = []
    max_map = {}

    for row in ws_props.iter_rows(min_row=2, values_only=True):
        pid = str(row[idx_id]) if row[idx_id] is not None else ""
        if not pid:
            continue

        # Tag handling
        newtag = row[idx_apply]
        newtag = str(newtag).strip() if newtag is not None else ""
        if newtag:
            if allowed_set and newtag not in allowed_set:
                invalid_tags.append(f"{pid}: {newtag}")
            tag_map[pid] = newtag

        # MaxChannels handling (optional)
        if idx_max is not None:
            mc_val = row[idx_max]
            if mc_val is not None and str(mc_val).strip() != "":
                try:
                    mc_int = int(mc_val)
                    max_map[pid] = mc_int
                except ValueError:
                    # Not a valid int; skip
                    pass

    return tag_map, invalid_tags, max_map


# ---------------------------------------------------------------------
# .lorprev helpers
# ---------------------------------------------------------------------


def check_preview_match(lorprev_path: Path, meta: dict):
    tree = ET.parse(lorprev_path)
    root = tree.getroot()
    if root.tag != "PreviewClass":
        raise ValueError(f"Unexpected root tag {root.tag!r}, expected 'PreviewClass'")

    preview_name = root.get("Name", "")
    preview_id = root.get("id", "")

    meta_stem = meta.get("SourceLorprevStem", "")
    meta_preview_name = meta.get("PreviewName", "")
    meta_preview_id = meta.get("PreviewId", "")

    errors = []

    if meta_stem and lorprev_path.stem != meta_stem:
        errors.append(f"File stem mismatch: Excel={meta_stem}, lorprev={lorprev_path.stem}")
    if meta_preview_name and preview_name != meta_preview_name:
        errors.append(f"PreviewName mismatch: Excel={meta_preview_name}, lorprev={preview_name}")
    if meta_preview_id and meta_preview_id != "" and preview_id != meta_preview_id:
        errors.append(f"PreviewId mismatch: Excel={meta_preview_id}, lorprev={preview_id}")

    return errors, tree, root


def apply_changes_to_lorprev(tree, root, tag_map, maxchannels_map):
    """
    Update PropClass attributes for all props with ids in the maps:
      - Tag from tag_map (PropID -> Tag string)
      - MaxChannels from maxchannels_map (PropID -> int)
    """
    props = root.findall("PropClass")
    updated_tags = 0
    updated_mc = 0

    for p in props:
        pid = p.get("id", "")
        if not pid:
            continue

        # Tags
        if pid in tag_map:
            new_tag = tag_map[pid]
            if p.get("Tag") != new_tag:
                p.set("Tag", new_tag)
                updated_tags += 1

        # MaxChannels
        if pid in maxchannels_map:
            new_mc = str(maxchannels_map[pid])
            if p.get("MaxChannels") != new_mc:
                p.set("MaxChannels", new_mc)
                updated_mc += 1

    return updated_tags, updated_mc


# ---------------------------------------------------------------------
# UI helper for picking Excel when no args given
# ---------------------------------------------------------------------


def pick_excel_file():
    root = tk.Tk()
    root.withdraw()
    excel_path = filedialog.askopenfilename(
        title="Select *_lor_groups.xlsx file",
        filetypes=[("Excel Workbook", "*.xlsx"), ("All Files", "*.*")],
    )
    if not excel_path:
        return None
    return Path(excel_path)


# ---------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(
        description=(
            "Apply ApplyTag / ApplyPropTag / MaxChannels values "
            "from *_lor_groups.xlsx to a .lorprev file.\n"
            "If --excel/--lorprev are omitted, a file picker is used for the Excel file and\n"
            "the .lorprev path is resolved from the Meta sheet."
        ),
        add_help=True,
    )
    parser.add_argument("--excel", help="Path to the *_lor_groups.xlsx file")
    parser.add_argument("--lorprev", help="Path to the corresponding .lorprev file")
    parser.add_argument(
        "--out",
        help=(
            "Output .lorprev path. "
            "Default (when omitted): <excel_folder>\\retagged_previews\\<SourceLorprevFileName>"
        ),
    )

    args = parser.parse_args()

    # Resolve Excel path
    if args.excel:
        xlsx_path = Path(args.excel)
    else:
        xlsx_path = pick_excel_file()
        if not xlsx_path:
            print("No Excel file selected. Exiting.")
            sys.exit(0)

    if not xlsx_path.is_file():
        print(f"Excel file not found: {xlsx_path}")
        sys.exit(1)

    # Load workbook and sheets
    wb = load_workbook(xlsx_path, data_only=True)
    if "Meta" not in wb.sheetnames:
        print("ERROR: Excel file has no Meta sheet; cannot verify source.")
        sys.exit(1)

    ws_meta = wb["Meta"]
    ws_gm = wb["GroupMembers"] if "GroupMembers" in wb.sheetnames else None
    ws_props = wb["Props"] if "Props" in wb.sheetnames else None

    if ws_gm is None and ws_props is None:
        print("ERROR: Excel file has neither GroupMembers nor Props sheet; nothing to apply.")
        sys.exit(1)

    meta = load_meta(ws_meta)

    # Resolve lorprev path:
    if args.lorprev:
        lorprev_path = Path(args.lorprev)
    else:
        # Prefer SourceLorprevFullPath from Meta if it exists
        full_path_str = meta.get("SourceLorprevFullPath", "")
        stem_name = meta.get("SourceLorprevStem", "")
        file_name = meta.get("SourceLorprevFileName", "")

        candidate = None
        if full_path_str:
            p = Path(full_path_str)
            if p.is_file():
                candidate = p

        if candidate is None and file_name:
            # Try Excel folder + SourceLorprevFileName
            candidate = xlsx_path.parent / file_name
            if not candidate.is_file():
                candidate = None

        if candidate is None and stem_name:
            # Last resort: Excel folder + <stem>.lorprev
            candidate = xlsx_path.parent / f"{stem_name}.lorprev"
            if not candidate.is_file():
                candidate = None

        if candidate is None:
            print("ERROR: Could not locate the source .lorprev file from Meta.")
            print("       Use --lorprev to specify it explicitly.")
            sys.exit(1)

        lorprev_path = candidate

    if not lorprev_path.is_file():
        print(f".lorprev file not found: {lorprev_path}")
        sys.exit(1)

    # Verify that Excel and .lorprev actually match (same preview)
    errors, tree, root = check_preview_match(lorprev_path, meta)
    if errors:
        print("ERROR: Excel workbook does not appear to match the selected .lorprev:")
        for e in errors:
            print("  -", e)
        print("Aborting without changes.")
        sys.exit(1)

    allowed_set = load_allowed_tags_from_lor(lorprev_path)
    if not allowed_set:
        print("[WARN] No previewtags.txt found (LOR path or preview folder).")
        print("       Proceeding WITHOUT tag validity checks.")

    # Collect changes from GroupMembers
    group_map, conflicts_group, invalid_group = ({}, [], [])
    if ws_gm is not None:
        group_map, conflicts_group, invalid_group = load_apply_map_groups(ws_gm, allowed_set)

    # Collect changes from Props (tags + maxchannels)
    prop_map, invalid_props, prop_max_map = ({}, [], {})
    if ws_props is not None:
        prop_map, invalid_props, prop_max_map = load_apply_map_props(ws_props, allowed_set)

    conflicts = list(conflicts_group)
    invalid_all = list(invalid_group) + list(invalid_props)

    if allowed_set and invalid_all:
        print("ERROR: Some ApplyTag/ApplyPropTag values are not in previewtags.txt:")
        for it in invalid_all:
            print("  -", it)
        print("Fix these tags in Excel (use the dropdown) and re-run. No changes applied.")
        sys.exit(1)

    # Merge group_map and prop_map -> merged_map (tags)
    merged_tag_map = {}
    for pid, tag in group_map.items():
        merged_tag_map[pid] = tag

    for pid, tag in prop_map.items():
        if pid in merged_tag_map and merged_tag_map[pid] != tag:
            conflicts.append(f"PropID {pid} tag mismatch: {merged_tag_map[pid]!r} vs {tag!r}")
        merged_tag_map[pid] = tag

    if conflicts:
        print("ERROR: Conflicting tag assignments found for the same PropID:")
        for c in conflicts:
            print("  -", c)
        print("Resolve conflicts in Excel and re-run. No changes applied.")
        sys.exit(1)

    if not merged_tag_map and not prop_max_map:
        print("No ApplyTag / ApplyPropTag / MaxChannels values found to apply. Nothing to do.")
        sys.exit(0)

    updated_tags, updated_mc = apply_changes_to_lorprev(tree, root, merged_tag_map, prop_max_map)

    # Determine output path:
    if args.out:
        out_path = Path(args.out)
    else:
        # Default: <excel_folder>\retagged_previews\<SourceLorprevFileName>
        out_dir = xlsx_path.parent / "retagged_previews"
        out_dir.mkdir(parents=True, exist_ok=True)
        file_name = meta.get("SourceLorprevFileName") or lorprev_path.name
        out_path = out_dir / file_name

    tree.write(out_path, encoding="utf-8", xml_declaration=True)

    print(f"[DONE] Updated {updated_tags} PropClass tags and {updated_mc} MaxChannels.")
    print(f"Output written to: {out_path}")


if __name__ == "__main__":
    main()
