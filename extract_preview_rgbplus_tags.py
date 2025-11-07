# GAL 2025-11-06  V1.3.0
# extract_preview_rgbplus_tags.py
#
# Light-O-Rama .lorprev Group/Tag Explorer → Excel
#
# - Reads a .lorprev preview
# - Builds:
#     * Groups        (one row per PropGroup)
#     * GroupMembers  (one row per Group+Prop member)
#     * Props         (one row per PropClass, grouped + ungrouped)
#     * TagsIndex     (unique tags in this preview)
#     * AllowedTags   (LOR master tag list from previewtags.txt)
#     * Meta          (ties Excel back to the source .lorprev)
# - Adds:
#     * GroupMembers.ApplyTag   (prefilled from MemberPropTag)
#     * Props.ApplyPropTag      (prefilled from PropTag)
# - Adds data validation dropdowns on:
#     * GroupMembers.ApplyTag
#     * Props.ApplyPropTag
#   using AllowedTags (or TagsIndex as fallback)
# - Copies C:\lor\LORInternal\Downloads\previews\previewtags.txt into
#   the preview folder as previewtags.txt (back door for other tools).
#
# Requires: pandas, openpyxl
#   pip install pandas openpyxl

import xml.etree.ElementTree as ET
from pathlib import Path
import pandas as pd
from collections import defaultdict
import tkinter as tk
from tkinter import filedialog, messagebox
import os
import subprocess
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ---------------------------------------------------------------------
# Tag source handling (LOR install + back door copy)
# ---------------------------------------------------------------------


def load_allowed_tags_from_lor(lorprev_path: Path):
    r"""
    Load allowed tags from the LOR-installed previewtags.txt.

    Primary:  C:\lor\LORInternal\Downloads\previews\previewtags.txt
    Back door: <lorprev_folder>\previewtags.txt

    Returns a list of strings (may be empty).
    """
    # Primary LOR path (raw string to avoid escape issues)
    default_path = Path(r"C:\lor\LORInternal\Downloads\previews\previewtags.txt")

    # Optional override via environment variable (for dev/testing)
    override = os.environ.get("LOR_PREVIEW_TAGS_PATH")
    tag_file = Path(override) if override else default_path

    tags = []

    if tag_file.is_file():
        # Read from LOR install
        with tag_file.open("r", encoding="utf-8") as f:
            for line in f:
                t = line.strip()
                if t:
                    tags.append(t)

        # Back door: copy to preview folder
        try:
            dest = lorprev_path.parent / "previewtags.txt"
            dest.write_text("\n".join(tags), encoding="utf-8")
        except Exception as e:
            print(f"[WARN] Could not copy previewtags.txt to preview folder: {e}")

        return tags

    # If LOR path is missing, try back door in preview folder
    fallback = lorprev_path.parent / "previewtags.txt"
    if fallback.is_file():
        with fallback.open("r", encoding="utf-8") as f:
            for line in f:
                t = line.strip()
                if t:
                    tags.append(t)
        return tags

    # Nothing found
    return []


# ---------------------------------------------------------------------
# Core parsing / DataFrame builders
# ---------------------------------------------------------------------


def parse_lorprev(path: Path):
    tree = ET.parse(path)
    root = tree.getroot()
    if root.tag != "PreviewClass":
        raise ValueError(f"Unexpected root tag {root.tag!r}, expected 'PreviewClass'")
    props = {p.get("id"): p for p in root.findall("PropClass")}
    groups = list(root.findall("PropGroup"))
    preview_name = root.get("Name", "")
    preview_id = root.get("id", "")
    return props, groups, preview_name, preview_id


def build_groups_df(groups):
    rows = []
    for g in sorted(groups, key=lambda g: g.get("Tag") or ""):
        rows.append(
            {
                "GroupTag": g.get("Tag", ""),
                "GroupName": g.get("Name", ""),
                "GroupID": g.get("id", ""),
                "MemberCount": len(g.findall("member")),
            }
        )
    return pd.DataFrame(rows)


def build_group_members_df(groups, props):
    rows = []
    for g in sorted(groups, key=lambda g: g.get("Tag") or ""):
        for m in g.findall("member"):
            pid = m.get("id")
            p = props.get(pid)
            if not p:
                continue
            rows.append(
                {
                    "GroupTag": g.get("Tag", ""),
                    "GroupName": g.get("Name", ""),
                    "GroupID": g.get("id", ""),
                    "MemberPropTag": p.get("Tag", ""),
                    "MemberPropName": p.get("Name", ""),
                    "MemberPropID": pid,
                    "ChannelGrid": p.get("ChannelGrid", ""),
                    "DeviceType": p.get("DeviceType", ""),
                    "MaxChannels": p.get("MaxChannels", ""),
                }
            )
    return pd.DataFrame(rows)


def build_props_df(props, groups):
    """Return a DataFrame of all PropClass elements, grouped + ungrouped."""
    # Map prop ID → list of groups that include it (for quick reference)
    membership = {}
    for g in groups:
        gname = g.get("Name", "")
        for m in g.findall("member"):
            pid = m.get("id")
            if not pid:
                continue
            membership.setdefault(pid, []).append(gname)

    rows = []
    for pid, p in props.items():
        rows.append(
            {
                "PropName": p.get("Name", ""),
                "PropTag": p.get("Tag", ""),
                "ApplyPropTag": p.get("Tag", ""),
                "ChannelGrid": p.get("ChannelGrid", ""),
                "DeviceType": p.get("DeviceType", ""),
                "MaxChannels": p.get("MaxChannels", ""),
                "PropID": pid,
                "Groups": ", ".join(membership.get(pid, [])),
            }
        )
    return pd.DataFrame(rows)


def build_tags_index_df(groups, props):
    tag_roles = defaultdict(set)
    for p in props.values():
        t = p.get("Tag") or ""
        if t:
            tag_roles[t].add("prop")
    for g in groups:
        t = g.get("Tag") or ""
        if t:
            tag_roles[t].add("group")

    rows = []
    for tag in sorted(tag_roles.keys()):
        roles = tag_roles[tag]
        rows.append(
            {
                "Tag": tag,
                "IsProp": "Y" if "prop" in roles else "",
                "IsGroup": "Y" if "group" in roles else "",
                "Role": (
                    "prop-only"
                    if roles == {"prop"}
                    else "group-only"
                    if roles == {"group"}
                    else "prop+group"
                ),
            }
        )
    return pd.DataFrame(rows)


def build_meta_df(lorprev_path: Path, preview_name: str, preview_id: str):
    rows = [
        {"Key": "SourceLorprevFileName", "Value": lorprev_path.name},
        {"Key": "SourceLorprevStem", "Value": lorprev_path.stem},
        {"Key": "SourceLorprevFullPath", "Value": str(lorprev_path)},
        {"Key": "PreviewName", "Value": preview_name},
        {"Key": "PreviewId", "Value": preview_id},
        {"Key": "ToolName", "Value": "extract_preview_rgbplus_tags"},
        {"Key": "ToolVersion", "Value": "1.1.0"},
    ]
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------
# UI helpers
# ---------------------------------------------------------------------


def pick_input_output():
    root = tk.Tk()
    root.withdraw()

    lorprev_path = filedialog.askopenfilename(
        title="Select Light-O-Rama .lorprev file",
        filetypes=[("LOR Preview Files", "*.lorprev"), ("All Files", "*.*")],
    )
    if not lorprev_path:
        return None, None

    default_name = Path(lorprev_path).stem + "_lor_groups.xlsx"
    out_path = filedialog.asksaveasfilename(
        title="Save Excel Workbook As",
        defaultextension=".xlsx",
        initialfile=default_name,
        filetypes=[("Excel Workbook", "*.xlsx")],
    )
    if not out_path:
        return lorprev_path, None

    return lorprev_path, out_path


def autofit(writer):
    """Auto-fit all columns (max width 60)."""
    for ws in writer.sheets.values():
        for col_cells in ws.columns:
            max_len = max(
                (len(str(c.value)) if c.value is not None else 0) for c in col_cells
            )
            ws.column_dimensions[col_cells[0].column_letter].width = min(
                max_len + 2, 60
            )


def add_applytag_validation(writer):
    """
    Add data validation to:
      - GroupMembers.ApplyTag
      - Props.ApplyPropTag

    Preference:
      1) AllowedTags.Tag (global LOR tag list)
      2) TagsIndex.Tag (fallback if AllowedTags is empty)
    """
    wb = writer.book
    ws_tags = writer.sheets.get("TagsIndex")
    ws_allowed = writer.sheets.get("AllowedTags")

    # Determine validation range
    dv_range = None
    if ws_allowed is not None and ws_allowed.max_row >= 2:
        max_row_allowed = ws_allowed.max_row
        dv_range = f"=AllowedTags!$A$2:$A${max_row_allowed}"
    elif ws_tags is not None and ws_tags.max_row >= 2:
        max_row_tags = ws_tags.max_row
        dv_range = f"=TagsIndex!$A$2:$A${max_row_tags}"

    if not dv_range:
        return

    # GroupMembers.ApplyTag
    ws_gm = writer.sheets.get("GroupMembers")
    if ws_gm is not None:
        apply_col_idx = None
        for cell in ws_gm[1]:
            if (cell.value or "").strip() == "ApplyTag":
                apply_col_idx = cell.column
                break
        if apply_col_idx is not None:
            apply_col_letter = get_column_letter(apply_col_idx)
            max_row_gm = ws_gm.max_row
            dv_gm = DataValidation(
                type="list",
                formula1=dv_range,
                allow_blank=True,
                showDropDown=True,
            )
            ws_gm.add_data_validation(dv_gm)
            dv_gm.add(f"{apply_col_letter}2:{apply_col_letter}{max_row_gm}")

    # Props.ApplyPropTag
    ws_props = writer.sheets.get("Props")
    if ws_props is not None:
        applyp_col_idx = None
        for cell in ws_props[1]:
            if (cell.value or "").strip() == "ApplyPropTag":
                applyp_col_idx = cell.column
                break
        if applyp_col_idx is not None:
            applyp_col_letter = get_column_letter(applyp_col_idx)
            max_row_props = ws_props.max_row
            dv_p = DataValidation(
                type="list",
                formula1=dv_range,
                allow_blank=True,
                showDropDown=True,
            )
            ws_props.add_data_validation(dv_p)
            dv_p.add(f"{applyp_col_letter}2:{applyp_col_letter}{max_row_props}")


def open_excel(path: Path):
    """Open the resulting Excel file using the default system app."""
    try:
        if os.name == "nt":  # Windows
            os.startfile(path)
        else:
            import sys

            subprocess.Popen(
                ["open" if sys.platform == "darwin" else "xdg-open", str(path)]
            )
    except Exception as e:
        print(f"[WARN] Could not open Excel automatically: {e}")


# ---------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------


def main():
    lorprev_file, out_file = pick_input_output()
    if not lorprev_file:
        print("No input selected.")
        return
    if not out_file:
        print("No output selected.")
        return

    lor_path = Path(lorprev_file)
    print(f"[INFO] Parsing: {lor_path.name}")

    try:
        props, groups, preview_name, preview_id = parse_lorprev(lor_path)
    except Exception as e:
        messagebox.showerror("Parse Error", f"Failed to parse {lor_path.name}\n\n{e}")
        return

    groups_df = build_groups_df(groups)
    gm_df = build_group_members_df(groups, props)
    props_df = build_props_df(props, groups)
    tags_df = build_tags_index_df(groups, props)
    meta_df = build_meta_df(lor_path, preview_name, preview_id)

    # Insert ApplyTag and prefill from MemberPropTag
    if not gm_df.empty and "MemberPropTag" in gm_df.columns:
        insert_idx = gm_df.columns.get_loc("MemberPropTag") + 1
        gm_df.insert(insert_idx, "ApplyTag", gm_df["MemberPropTag"])
    else:
        gm_df["ApplyTag"] = ""

    # Reorder GroupMembers columns to a human-friendly layout
    desired_order = [
        "GroupTag",
        "GroupName",
        "MemberPropName",
        "MemberPropTag",
        "ApplyTag",
        "ChannelGrid",
        "DeviceType",
        "MaxChannels",
        "GroupID",
        "MemberPropID",
    ]
    existing = list(gm_df.columns)
    ordered = [c for c in desired_order if c in existing]
    extra = [c for c in existing if c not in ordered]
    if ordered:
        gm_df = gm_df[ordered + extra]

    # Load allowed tags from LOR / preview folder
    allowed_tags = load_allowed_tags_from_lor(lor_path)
    if allowed_tags:
        allowed_df = pd.DataFrame({"Tag": allowed_tags})
    else:
        allowed_df = pd.DataFrame(columns=["Tag"])

    out_path = Path(out_file)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        groups_df.to_excel(writer, index=False, sheet_name="Groups")
        gm_df.to_excel(writer, index=False, sheet_name="GroupMembers")
        props_df.to_excel(writer, index=False, sheet_name="Props")
        tags_df.to_excel(writer, index=False, sheet_name="TagsIndex")
        meta_df.to_excel(writer, index=False, sheet_name="Meta")
        allowed_df.to_excel(writer, index=False, sheet_name="AllowedTags")

        autofit(writer)
        add_applytag_validation(writer)

    messagebox.showinfo("Export Complete", f"Excel workbook created:\n{out_path}")
    print(f"[DONE] Excel written: {out_path}")

    open_excel(out_path)


if __name__ == "__main__":
    main()
