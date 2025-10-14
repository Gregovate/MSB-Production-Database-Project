#!/usr/bin/env python3
r"""
Compare LOR v6 DB display names to the Google Sheet 'Displays' export and
write a formatted Excel report.

Default paths are the same style as parse_props_v6.py (prompt with defaults).

Usage:
    python compare_displays_vs_db.py
    # or override any/all via CLI:
    python compare_displays_vs_db.py "<db_path>" "<csv_path>" "<xlsx_path>"

Dependencies:
    pip install pandas openpyxl
"""

import os
import re
import sys
import sqlite3
import datetime
from pathlib import Path
import pandas as pd
import csv

# ============================= G: ONLY ============================= #
G = Path(r"G:\Shared drives\MSB Database")

def require_g():
    if not G.exists():
        print("[FATAL] G: drive not available. All data lives on the shared drive.")
        print("        Mount the shared drive and try again.")
        sys.exit(2)

# Hard-set defaults (always on G:\)
DEFAULT_DB_PATH   = G / "database"    / "lor_output_v6.db"
DEFAULT_CSV_PATH  = G / "Spreadsheet" / "displays_export.csv"
DEFAULT_XLSX_PATH = G / "Spreadsheet" / "lor_display_compare.xlsx"

print(f"[INFO] Default DB   : {DEFAULT_DB_PATH}")
print(f"[INFO] Default CSV  : {DEFAULT_CSV_PATH}")
print(f"[INFO] Default XLSX : {DEFAULT_XLSX_PATH}")


DB_DISPLAY_COL = "LORComment"
USE_FUZZY = True


# --------- small helpers ---------
def get_path(prompt: str, default_path: Path) -> Path:
    """Prompt for a path, using a default if user just hits Enter."""
    raw = input(f"{prompt} [{default_path}]: ").strip()
    if raw:
        # Normalize backslashes etc. to a clean Path
        return Path(os.path.normpath(raw))
    return default_path

def norm(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"\s+", " ", s)
    return s.lower()

def levenshtein(a: str, b: str) -> int:
    if a == b:
        return 0
    if not a:
        return len(b)
    if not b:
        return len(a)
    prev = list(range(len(b) + 1))
    curr = [0] * (len(b) + 1)
    for i, ca in enumerate(a, 1):
        curr[0] = i
        for j, cb in enumerate(b, 1):
            cost = 0 if ca == cb else 1
            curr[j] = min(curr[j-1] + 1, prev[j] + 1, prev[j-1] + cost)
        prev, curr = curr, prev
    return prev[-1]

def norm_key(s: str) -> str:
    s = (s or "").strip().lower().replace('\u00a0', ' ')
    s = re.sub(r'\b(left|lh)\b', 'l', s)
    s = re.sub(r'\b(right|rh)\b', 'r', s)
    s = re.sub(r'[^a-z0-9]+', '', s)          # remove dashes/spaces/etc.
    s = re.sub(r'\d+', lambda m: str(int(m.group())), s)  # 01 -> 1
    return s

# Clean but preserve exact spelling/case (only trim NBSP and outer spaces)
def exact_clean(s: str) -> str:
    return (s or "").replace('\u00a0', ' ').strip()

_reason_tests = [
    ("dash_vs_space",      lambda a,b: ("-" in a) ^ ("-" in b) or (" " in a) ^ (" " in b)),
    ("zero_pad_diff",      lambda a,b: re.sub(r'\d+', lambda m: str(int(m.group())), a) == re.sub(r'\d+', lambda m: str(int(m.group())), b) and a != b),
    ("lh_rh_vs_l_r",       lambda a,b: re.sub(r'\b(lh|rh)\b', lambda m: m.group(0)[0], a.lower()) == re.sub(r'\b(lh|rh)\b', lambda m: m.group(0)[0], b.lower()) and a != b),
    ("case_only",          lambda a,b: a.lower() == b.lower() and a != b),
]

def diff_reason(a: str, b: str) -> str:
    a = a or ""; b = b or ""
    hits = [name for name,fn in _reason_tests if fn(a,b)]
    return ", ".join(hits) if hits else ""

# ------ Cleanup Duplicate Headers ------ 25/10/14 GAL

CSV_BASENAME = "displays_export.csv"

def normalize_csv_path(p: Path) -> Path:
    """
    If the user passes a folder instead of a file, point to ...\\displays_export.csv.
    Accepts paths with a trailing backslash.
    """
    try:
        # If it's an existing directory, use default basename inside it
        if p.exists() and p.is_dir():
            return p / CSV_BASENAME
        # Handle trailing slash case that still resolves to a directory
        s = str(p).rstrip()
        if s.endswith(os.sep):
            pp = Path(s)
            if pp.exists() and pp.is_dir():
                return pp / CSV_BASENAME
    except Exception:
        pass
    return p

def _normalize_header_names(header_cells):
    """Preserve given names; synthesize placeholders for blank headers to keep positions."""
    out, used = [], set()
    for i, raw in enumerate(header_cells):
        name = (raw or "").strip()
        if not name:
            name = f"__col{i+1}"
        base = name
        n = 2
        while name in used:
            name = f"{base}__{n}"
            n += 1
        used.add(name)
        out.append(name)
    return out

def _find_year_idx(headers):
    # accept "YearBuilt" or "Year Built"
    norm = [re.sub(r"\s+", "", h or "").lower() for h in headers]
    for want in ("yearbuilt", "year"):
        if want in norm:
            return norm.index(want)
    return None

def _looks_int(s: str) -> bool:
    s = (s or "").strip()
    return bool(re.fullmatch(r"-?\d{1,6}", s))



def clean_duplicate_header_inplace(csv_path: Path) -> None:
    """
    If row 2 repeats the header (or there are blank lines before the header),
    rewrite the CSV so there's a single header on row 1 and data starts on row 2.
    Safe for UTF-8 with BOM and Windows newlines.
    """
    if not csv_path.exists():
        return

    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))

    # Find the first non-empty row as the header
    i = 0
    while i < len(rows) and not any((c or "").strip() for c in rows[i]):
        i += 1
    if i >= len(rows):
        return  # empty file, nothing to do

    header = [(c or "").strip() for c in rows[i]]

    # Advance to the first candidate data row
    j = i + 1
    # Skip blank lines
    while j < len(rows) and not any((c or "").strip() for c in rows[j]):
        j += 1
    # If the very next non-blank row equals the header, skip it
    if j < len(rows) and [(c or "").strip() for c in rows[j]] == header:
        j += 1

    cleaned = [header]
    for r in rows[j:]:
        if not r or not any((c or "").strip() for c in r):
            continue
        cleaned.append(r)

    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerows(cleaned)

# ------ Wait for Sync to finish ------ 25/10/14 GAL
# add just below your other small helpers / before wait_until_stable
from pathlib import Path
import time, hashlib

def wait_until_stable(path: Path, stable_secs=3, timeout=120):
    start = time.time()
    last = None; last_change = time.time()
    while True:
        if path.exists():
            try:
                sig = (path.stat().st_size, hashlib.sha256(path.read_bytes()).hexdigest())
            except Exception:
                sig = None
            if sig and sig == last:
                if time.time() - last_change >= stable_secs:
                    return
            else:
                last = sig; last_change = time.time()
        if time.time() - start > timeout:
            raise TimeoutError(f"{path} not stable after {timeout}s")
        time.sleep(0.75)

# --------- loaders ---------
def load_db_names(db_path: Path) -> pd.DataFrame:
    conn = sqlite3.connect(db_path)
    try:
        q = f"""
            SELECT
                p.{DB_DISPLAY_COL} AS display_name,
                pv.Name            AS Preview,
                pv.Revision        AS Revision
            FROM props p
            JOIN previews pv ON pv.id = p.PreviewId
            WHERE TRIM(COALESCE(p.{DB_DISPLAY_COL}, '')) <> ''
        """
        df = pd.read_sql_query(q, conn)
    finally:
        conn.close()

    df["display_name"]      = df["display_name"].astype(str)
    df["display_name_clean"] = df["display_name"].map(exact_clean)
    df["display_name_norm"]  = df["display_name"].map(norm_key)
    df = df[df["display_name_norm"] != ""].drop_duplicates(subset=["display_name_clean"])
    return df

def load_sheet_names(csv_path: Path) -> pd.DataFrame:
    """
    Robust loader for the Displays export:
      - UTF-8 BOM safe
      - First non-blank row is header (blank header cells allowed)
      - Second row is skipped if YearBuilt isn't numeric (human labels row)
      - '#N/A' normalized to ''
    Returns DataFrame with columns: display_name, display_name_clean, display_name_norm
    """
    print(f"[INFO] Reading sheet CSV from: {csv_path}")

    # Parse with csv module (more tolerant than pandas for ragged rows)
    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        r = csv.reader(f, delimiter=",")

        # Find header
        for row in r:
            if row and any((c or "").strip() for c in row):
                raw_header = [ (c or "").strip() for c in row ]
                break
        else:
            return pd.DataFrame(columns=["display_name","display_name_clean","display_name_norm"])

        headers = _normalize_header_names(raw_header)
        year_idx = _find_year_idx(headers)

        # Identify display-name column by header
        # Accept 'DisplayName', 'Display Name', 'Display', or 'LOR Comment'
        disp_idx = None
        norm_hdrs = [ (h or "").strip().lower().replace("_"," ") for h in headers ]
        for want in ("display name", "displayname", "display", "lor comment", "lorcomment"):
            if want in norm_hdrs:
                disp_idx = norm_hdrs.index(want)
                break
        if disp_idx is None:
            disp_idx = 0  # fallback to first column

        # Consume rows
        disp_values = []
        checked_second = False
        for row in r:
            if not row or not any((c or "").strip() for c in row):
                continue
            cells = [ (c or "").strip() for c in row ]
            # pad/truncate to header length
            if len(cells) < len(headers):
                cells += [""] * (len(headers) - len(cells))
            elif len(cells) > len(headers):
                cells = cells[:len(headers)]

            # Skip labels row if YearBuilt not numeric
            if not checked_second:
                checked_second = True
                if year_idx is not None and not _looks_int(cells[year_idx]):
                    continue

            v = cells[disp_idx]
            if v.upper() == "#N/A":
                v = ""
            disp_values.append(v)

    df = pd.DataFrame({"display_name": disp_values})
    df["display_name"]       = df["display_name"].astype(str)
    df["display_name_clean"] = df["display_name"].map(exact_clean)
    df["display_name_norm"]  = df["display_name"].map(norm_key)
    df = df[df["display_name_norm"] != ""].drop_duplicates(subset=["display_name_clean"])

    print(f"[INFO] Using column for display name: '{headers[disp_idx]}'")
    return df


def fuzzy_suggestions(missing_df: pd.DataFrame, universe_df: pd.DataFrame, k=3) -> pd.DataFrame:
    if missing_df.empty or universe_df.empty:
        return pd.DataFrame(columns=["Missing_Side", "From_Name", "Suggested_Match", "Distance"])

    miss = missing_df["display_name_norm"].tolist()
    uni  = universe_df["display_name_norm"].tolist()

    rows = []
    for key in miss:
        ranked = sorted(((levenshtein(key, u), u) for u in uni), key=lambda x: x[0])[:k]
        for d, u in ranked:
            rows.append((key, u, d))

    out = pd.DataFrame(rows, columns=["From_Key","To_Key","Distance"])

    key_to_name = pd.concat([
        missing_df[["display_name_norm","display_name"]],
        universe_df[["display_name_norm","display_name"]],
    ]).drop_duplicates(subset=["display_name_norm"]).set_index("display_name_norm")["display_name"].to_dict()

    out["From_Name"] = out["From_Key"].map(lambda k: key_to_name.get(k, k))
    out["Suggested_Match"] = out["To_Key"].map(lambda k: key_to_name.get(k, k))
    out = out.drop(columns=["From_Key","To_Key"]).sort_values(["Distance","From_Name","Suggested_Match"])
    return out


# --------- compare ---------
def run_compare(db_path: Path, csv_path: Path) -> dict:
    if not db_path.exists():
        raise FileNotFoundError(f"DB not found: {db_path}")
    if not csv_path.exists():
        raise FileNotFoundError(f"CSV not found: {csv_path}\n(Hint: export from Sheets first)")

    db_df = load_db_names(db_path)
    sh_df = load_sheet_names(csv_path)

    # Sanity probes
    for probe in ["ClarksWagon","WallyWagon"]:
        in_sheet = (sh_df["display_name"].str.contains(probe, case=False, na=False)).any()
        in_db    = (db_df["display_name"].str.contains(probe, case=False, na=False)).any()
        print(f"[CHECK] {probe}: sheet={in_sheet} db={in_db}")

    # 1) EXACT matches (contract) — strict equality after only NBSP/trim cleanup
    matches = db_df.merge(
        sh_df,
        left_on="display_name_clean",
        right_on="display_name_clean",
        suffixes=("_db","_sheet")
    ).rename(columns={
        "display_name_db": "DB_Display_Name",
        "display_name_sheet": "Sheet_Display_Name",
        "display_name_clean":"Exact_Key",
        "display_name_norm_db":"DB_Normalized",
        "display_name_norm_sheet":"Sheet_Normalized",
    })[["DB_Display_Name","Sheet_Display_Name","Exact_Key","Preview","Revision"]].sort_values("Exact_Key")

    # 2) DB_only / Sheet_only (still strict)
    db_only = db_df[~db_df["display_name_clean"].isin(matches["Exact_Key"])] \
              .rename(columns={"display_name":"DB_Display_Name","display_name_norm":"Normalized_Key"}) \
              [["DB_Display_Name","display_name_clean","Normalized_Key","Preview","Revision"]] \
              .sort_values("DB_Display_Name").rename(columns={"display_name_clean":"Exact_Key"})

    sheet_only = sh_df[~sh_df["display_name_clean"].isin(matches["Exact_Key"])] \
                .rename(columns={"display_name":"Sheet_Display_Name","display_name_norm":"Normalized_Key"}) \
                [["Sheet_Display_Name","display_name_clean","Normalized_Key"]] \
                .sort_values("Sheet_Display_Name").rename(columns={"display_name_clean":"Exact_Key"})

    # 3) NEEDS_FIX — “intended pairs” that don’t match exactly (same norm_key, different strings)
    fix_pairs = db_only.merge(
        sheet_only,
        on="Normalized_Key",
        how="inner",
        suffixes=("_DBonly","_Sheetonly")
    )

    if not fix_pairs.empty:
        fix_pairs["Reason"] = fix_pairs.apply(
            lambda r: diff_reason(r["DB_Display_Name"], r["Sheet_Display_Name"]),
            axis=1
        )
        # Choose your canonical side here (Sheet or DB). Default: SHEET as authority.
        fix_pairs.insert(0, "Edit_Side", "Preview_Comment")  # or "Sheet"
        fix_pairs.insert(1, "Target_Name", fix_pairs["Sheet_Display_Name"])
        needs_fix = fix_pairs[[
            "Edit_Side","DB_Display_Name","Sheet_Display_Name","Target_Name","Reason","Normalized_Key"
        ]].sort_values(["Reason","Normalized_Key","Target_Name"])
    else:
        needs_fix = pd.DataFrame(columns=[
            "Edit_Side","DB_Display_Name","Sheet_Display_Name","Target_Name","Reason","Normalized_Key"
        ])

    # 4) Optional fuzzy suggestions (unchanged)
    near_matches = pd.DataFrame(columns=["Missing_Side","From_Name","Suggested_Match","Distance"])
    if USE_FUZZY:
        sug_db = fuzzy_suggestions(
            db_only.rename(columns={"DB_Display_Name":"display_name","Exact_Key":"display_name_norm"}),
            pd.concat([
                sheet_only.rename(columns={"Sheet_Display_Name":"display_name","Exact_Key":"display_name_norm"}),
                matches.rename(columns={"Sheet_Display_Name":"display_name","Exact_Key":"display_name_norm"})[["display_name","display_name_norm"]],
            ], ignore_index=True),
        )
        sug_db.insert(0, "Missing_Side", "DB_only")

        sug_sh = fuzzy_suggestions(
            sheet_only.rename(columns={"Sheet_Display_Name":"display_name","Exact_Key":"display_name_norm"}),
            pd.concat([
                db_only.rename(columns={"DB_Display_Name":"display_name","Exact_Key":"display_name_norm"}),
                matches.rename(columns={"DB_Display_Name":"display_name","Exact_Key":"display_name_norm"})[["display_name","display_name_norm"]],
            ], ignore_index=True),
        )
        sug_sh.insert(0, "Missing_Side", "Sheet_only")

        near_matches = pd.concat([sug_db, sug_sh], ignore_index=True)

    summary = pd.DataFrame([{
        "As of": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
        "DB total (unique)": len(db_df),
        "Sheet total (unique)": len(sh_df),
        "Exact matches": len(matches),
        "DB-only (strict)": len(db_only),
        "Sheet-only (strict)": len(sheet_only),
        "Needs_Fix (same intent, not exact)": len(needs_fix),
    }])

    return {
        "summary": summary,
        "matches": matches,
        "db_only": db_only,
        "sheet_only": sheet_only,
        "needs_fix": needs_fix,
        "near_matches": near_matches,
    }


# --------- Excel formatting ---------
def autosize_and_filter(ws):
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    from openpyxl.utils import get_column_letter
    for col in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value) if cell.value is not None else ""
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 80)

def write_excel(out_path: Path, tables: dict):
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for name in ["Summary","Matches","DB_only","Sheet_only","Needs_Fix","Near_matches"]:
            tables[name.lower()].to_excel(writer, sheet_name=name, index=False)

        wb = writer.book
        for sheet in wb.sheetnames:
            autosize_and_filter(wb[sheet])
    print(f"[OK] Wrote Excel report: {out_path}")



# --------- entrypoint ---------
def main():
    require_g()  # fail fast if shared drive isn't mounted

    # CLI overrides take precedence; otherwise prompt with G:\ defaults
    args = sys.argv[1:]
    if len(args) >= 1:
        db_path = Path(args[0])
    else:
        db_path = get_path("Enter database path", DEFAULT_DB_PATH)

    if len(args) >= 2:
        # user might paste the folder; normalize to ...\displays_export.csv
        csv_path = normalize_csv_path(Path(args[1]))
    else:
        # wording fix: we want the FILE, not just folder; still accept folder via normalize
        csv_input = get_path("Enter the path to Displays CSV (or its folder)", DEFAULT_CSV_PATH)
        csv_path = normalize_csv_path(Path(csv_input))

    if len(args) >= 3:
        xlsx_path = Path(args[2])
    else:
        xlsx_path = get_path("Enter output Excel path", DEFAULT_XLSX_PATH)

    # Enforce G:\ for all 3
    def _is_on_g(p: Path) -> bool:
        drv = getattr(p, "drive", "")
        return (drv.upper() == "G:") or str(p)[:2].upper() == "G:"

    for label, p in [("DB path", db_path), ("CSV path", csv_path), ("XLSX path", xlsx_path)]:
        if not _is_on_g(p):
            print(f"[FATAL] {label} must be on G:\\ — got: {p}")
            sys.exit(2)

    print(f"[INFO] Using database: {db_path}")
    print(f"[INFO] Using sheet CSV: {csv_path}")
    print(f"[INFO] Will write Excel: {xlsx_path}")

    # --- NEW: wait for Drive to finish and normalize the CSV header ---
    try:
        print(f"[WAIT] Ensuring CSV is present/stable: {csv_path}")
        wait_until_stable(csv_path, stable_secs=3, timeout=180)
        clean_duplicate_header_inplace(csv_path)
        print("[OK] CSV ready.")
    except Exception as e:
        print(f"[ERROR] CSV not ready: {e}")
        sys.exit(1)

    try:
        tables = run_compare(db_path, csv_path)
        write_excel(xlsx_path, tables)
    except Exception as e:
        print(f"[ERROR] {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()